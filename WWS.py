# -*- coding: utf-8 -*-
"""
Created on Mon Mar  6 17:11:58 2017

@author: Kumar.Singh
"""

import pandas as pd
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook
import os
import win32com.client as win32
from datetime import datetime

os.chdir('C:\\Users\\kumar.singh\\Desktop\\sharepoint')

#spLink = r"C:\Users\kumar.singh\Downloads\muQ status_03032017 (1).xlsx"
spLink = '''https://musigma.sharepoint.com/sites/Microsoft%20_WWS_Team/Shared%20Documents/04%20Daily%20Scrum/MS_WWS_Scrum.xlsx?web=1'''
saveTo = r'C:\Users\kumar.singh\Desktop\sharepoint\WWS.xlsx'

def spfetcher(spLink, saveTo):
    print('fetching data from SP')
    xl = win32.Dispatch("Excel.Application")
    wb = xl.Workbooks.Open(spLink)
    wb.SaveAs(saveTo) #, FileFormat = -4143 )
    wb.Close()
    xl.Quit()
    df = pd.read_excel(saveTo)
#    
    return(df)

def mailer( body , to):
    print('sending mail to:' , to)
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = to
    mail.Subject = 'Deliverables'
    mail.HTMLBody = body       # this field is optional 
    mail.Send()

#t_date = a.strftime('%B')+ ' ' + str(a.day)
#y_date = a.strftime('%B')+ ' ' + str((datetime.now() - BDay(1)).day)
##re.search('%s \d.*' %(a.strftime('%B')), b , re.I)
#t_scrum = df.ix[df.Date.str.contains('%s' %(t_date)), 1:]
#y_scrum = df.ix[df.Date.str.contains('%s' %(y_date)), 1:]
#y_scrum.to_html('y_scrum.htm', index = False ,  bor )
#t_scrum.to_html('t_scrum.htm', index = False )
            

def table_fetch(day, df, c):   
    df = df
    b = c
    if day == 'today':
        day = 1
    elif day == 'yesterday' :
        day = 2    
    wb = load_workbook('WWS.xlsx', data_only=True)
    sh = wb[wb.get_sheet_names()[0]]       
              
    for index, row in enumerate(sh.iter_rows()):
        a = list(set(df.Date))
        a.sort()
        t_date = df.Date[df.Date == a[-day]].index
        if (index > t_date.min()+1  and index < t_date.max() +3) or index == 0:
            print(index)
            for cell in row:
                a = "" if cell.value is  None else cell.value
                if cell.fill.start_color.index != '00000000':
                    b = b + '<td style = "background:#%s"> %s </td>'%(str(cell.fill.start_color.index)[2:8], a)
                else:
                    b = b + "<td > %s </td>"%( a)
            b +=r"</tr>"
        open('html.html', 'w').write(b)
    return(b)

def snap_mail():
    c = '''<style>
              table {
                border-collapse: collapse;
              }
              th, td {
                border: 1px solid black;
                padding: 10px;
                text-align: left;
                border-collapse:collapse;
              }
            </style><table> <tr>'''
    df = spfetcher(spLink, saveTo)
#    cs = pd.read_excel('WWS.xlsx', 1)
    to = 'MSFT_WWS_Leads@mu-sigma.com'
  #  to = 'kumar.singh@mu-sigma.com'
    mailer( table_fetch('today', df, c), to = to    )
    mailer( table_fetch('yesterday', df, c), to = to)
    os.remove(saveTo)








